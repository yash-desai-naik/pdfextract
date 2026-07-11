"""Excel (XLSX) report generator using openpyxl."""

from __future__ import annotations

from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

from src.utils.logging import get_logger
from src.models.rooms import Room

logger = get_logger("report.xlsx")


class XLSXReport:
    """Generates a formatted Excel report of the heating takeoff."""

    HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    TOTAL_FONT = Font(name="Calibri", bold=True, size=11)
    TOTAL_FILL = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
    DATA_FONT = Font(name="Calibri", size=10)
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    HEADERS = [
        "Room",
        "Measurements Used",
        "Gross Area (m²)",
        "Excluded Area (m²)",
        "Setback (m²)",
        "Net Area (m²)",
        "Strip Count",
        "Linear Metres",
        "Mat Area (m²)",
        "Coverage (%)",
        "Confidence",
    ]

    COL_WIDTHS = [20, 18, 14, 16, 12, 14, 12, 14, 14, 14, 12]

    def generate(self, rooms: list[Room], totals: dict, output_path: Path) -> Path:
        """Generate a formatted Excel workbook.

        Args:
            rooms: All processed rooms.
            totals: Project-wide totals.
            output_path: Destination .xlsx path.

        Returns:
            Path to the written file.
        """
        wb = Workbook()

        # --- Sheet 1: Room-by-room breakdown ---
        ws = wb.active
        ws.title = "Heating Takeoff"

        # Headers
        for col_idx, header in enumerate(self.HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.THIN_BORDER

        # Column widths
        for col_idx, width in enumerate(self.COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Data rows
        for row_idx, room in enumerate(rooms, 2):
            data = [
                room.name,
                room.measurements_used or "calculated",
                round(room.gross_area_m2, 3),
                round(room.excluded_area_m2, 3),
                round(room.setback_area_m2, 3),
                round(room.net_heatable_area_m2, 3),
                room.strip_count,
                round(room.total_linear_m, 3),
                round(room.mat_area_m2, 3),
                round(room.coverage_pct, 1),
                round(room.confidence, 2),
            ]
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.DATA_FONT
                cell.border = self.THIN_BORDER
                if col_idx >= 3:
                    cell.alignment = Alignment(horizontal="right")

        # Totals row
        total_row = len(rooms) + 2
        total_data = [
            "TOTAL",
            "",
            round(totals.get("total_gross_area_m2", 0), 3),
            round(totals.get("total_excluded_area_m2", 0), 3),
            round(totals.get("total_setback_area_m2", 0), 3),
            round(totals.get("total_net_heatable_area_m2", 0), 3),
            totals.get("total_strips", 0),
            round(totals.get("total_linear_m", 0), 3),
            round(totals.get("total_mat_area_m2", 0), 3),
            "",
            "",
        ]
        for col_idx, value in enumerate(total_data, 1):
            cell = ws.cell(row=total_row, column=col_idx, value=value)
            cell.font = self.TOTAL_FONT
            cell.fill = self.TOTAL_FILL
            cell.border = self.THIN_BORDER

        # --- Sheet 2: Summary ---
        ws2 = wb.create_sheet("Summary")
        summary_items = [
            ("Total Gross Area (m²)", round(totals.get("total_gross_area_m2", 0), 3)),
            ("Total Excluded Area (m²)", round(totals.get("total_excluded_area_m2", 0), 3)),
            ("Total Setback Area (m²)", round(totals.get("total_setback_area_m2", 0), 3)),
            ("Total Net Heatable Area (m²)", round(totals.get("total_net_heatable_area_m2", 0), 3)),
            ("Total Mat Area (m²)", round(totals.get("total_mat_area_m2", 0), 3)),
            ("Total Linear Metres", round(totals.get("total_linear_m", 0), 3)),
            ("Total Strips", totals.get("total_strips", 0)),
            ("Number of Rooms", totals.get("room_count", 0)),
        ]
        for row_idx, (label, value) in enumerate(summary_items, 1):
            ws2.cell(row=row_idx, column=1, value=label).font = Font(bold=True, size=11)
            ws2.cell(row=row_idx, column=2, value=value).font = self.DATA_FONT
        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 15

        # --- Sheet 3: Exclusion Details ---
        ws3 = wb.create_sheet("Exclusions")
        ws3.cell(row=1, column=1, value="Room").font = self.HEADER_FONT
        ws3.cell(row=1, column=1).fill = self.HEADER_FILL
        ws3.cell(row=1, column=2, value="Reason").font = self.HEADER_FONT
        ws3.cell(row=1, column=2).fill = self.HEADER_FILL
        ws3.cell(row=1, column=3, value="Area (m²)").font = self.HEADER_FONT
        ws3.cell(row=1, column=3).fill = self.HEADER_FILL
        ws3.cell(row=1, column=4, value="Source Type").font = self.HEADER_FONT
        ws3.cell(row=1, column=4).fill = self.HEADER_FILL
        ws3.column_dimensions["A"].width = 20
        ws3.column_dimensions["B"].width = 25
        ws3.column_dimensions["C"].width = 14
        ws3.column_dimensions["D"].width = 14

        row_idx = 2
        for room in rooms:
            for exc in room.exclusions:
                ws3.cell(row=row_idx, column=1, value=room.name).font = self.DATA_FONT
                ws3.cell(row=row_idx, column=2, value=exc.reason).font = self.DATA_FONT
                ws3.cell(row=row_idx, column=3, value=round(exc.area_m2, 3)).font = self.DATA_FONT
                ws3.cell(row=row_idx, column=4, value=exc.source_type).font = self.DATA_FONT
                row_idx += 1

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        logger.info("XLSX report written to %s", output_path)
        return output_path
